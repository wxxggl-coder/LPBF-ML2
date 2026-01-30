import pandas as pd
from sklearn.ensemble import ExtraTreesRegressor
from sklearn.model_selection import train_test_split
from sklearn.metrics import r2_score
import numpy as np
import math
from sklearn.metrics import mean_squared_error
import openpyxl
from sklearn.preprocessing import StandardScaler

# 读取数据
data = pd.read_csv('11.csv')




# 提取特征和目标值
X = data.iloc[:, :4].values
y = data.iloc[:, -1].values


model = ExtraTreesRegressor()


# 将数据分为训练集和测试集
x_train_true, x_test_true, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# 训练集中对应的x真实值
x_Ti_wt_train = np.array(x_train_true[:, 0])
x_B_wt_train = np.array(x_train_true[:, 1])
x_w_train = np.array(x_train_true[:, 2])
x_v_train = np.array(x_train_true[:, 3])
# 测试集中对应的x真实值
x_Ti_wt_test = np.array(x_test_true[:, 0])
x_B_wt_test = np.array(x_test_true[:, 1])
x_w_test = np.array(x_test_true[:, 2])
x_v_test = np.array(x_test_true[:, 3])

# 数据归一化
scaler = StandardScaler()
x_train = scaler.fit_transform(x_train_true)
x_test = scaler.transform(x_test_true)
model.fit(x_train, y_train)
# 在测试集上进行预测
y_pred_train = model.predict(x_train)
y_pred_test = model.predict(x_test)


# 计算预测值的准确度（R2分数）和均方根误差
accuracy = r2_score(y_test, y_pred_test)
sqrterror1 = math.sqrt(mean_squared_error(y_test, y_pred_test))
S_accuracy = 1 - np.mean(np.abs(y_pred_test - y_test) / y_test)


print("平均强度R2：", accuracy)
print("平均强度的均方根误差：", sqrterror1)
print('平均强度的预测准确度：', S_accuracy)

# 创建一个新的工作簿和工作表
wb = openpyxl.Workbook()
sheet = wb.active

# 写入表头
sheet.cell(row=1, column=1).value = "Ti质量分数(wt%)"
sheet.cell(row=1, column=2).value = "B质量分数(wt%)"
sheet.cell(row=1, column=3).value = "扫描功率(W)"
sheet.cell(row=1, column=4).value = "扫描速度(mm/s)"
sheet.cell(row=1, column=5).value = "真实塑性(MPa)"
sheet.cell(row=1, column=6).value = "预测塑性(MPa)"

# 写入数据
for i in range(1, len(x_train)+1):
    sheet.cell(row=i+1, column=1).value = float(x_Ti_wt_train[i-1])
    sheet.cell(row=i+1, column=2).value = float(x_B_wt_train[i-1])
    sheet.cell(row=i+1, column=3).value = float(x_w_train[i-1])
    sheet.cell(row=i+1, column=4).value = float(x_v_train[i-1])
    sheet.cell(row=i+1, column=5).value = float(y_train[i-1])
    sheet.cell(row=i+1, column=6).value = float(y_pred_train[i-1])

# 保存为.xlsx文件
wb.save('F:/Ti、B有能量密度/prediction_extra_tree_11_train_plasticity.xlsx')

# 创建一个新的工作簿和工作表
wb = openpyxl.Workbook()
sheet = wb.active

# 写入表头
sheet.cell(row=1, column=1).value = "Ti质量分数(wt%)"
sheet.cell(row=1, column=2).value = "B质量分数(wt%)"
sheet.cell(row=1, column=3).value = "扫描功率(W)"
sheet.cell(row=1, column=4).value = "扫描速度(mm/s)"
sheet.cell(row=1, column=5).value = "真实塑性(MPa)"
sheet.cell(row=1, column=6).value = "预测塑性(MPa)"

# 写入数据
for i in range(1, len(x_test)+1):
    sheet.cell(row=i+1, column=1).value = float(x_Ti_wt_test[i-1])
    sheet.cell(row=i+1, column=2).value = float(x_B_wt_test[i-1])
    sheet.cell(row=i+1, column=3).value = float(x_w_test[i-1])
    sheet.cell(row=i+1, column=4).value = float(x_v_test[i-1])
    sheet.cell(row=i+1, column=5).value = float(y_test[i - 1])
    sheet.cell(row=i+1, column=6).value = float(y_pred_test[i-1])

# 保存为.xlsx文件
wb.save('F:/Ti、B有能量密度/prediction_extra_tree_11_test_plasticity.xlsx')
